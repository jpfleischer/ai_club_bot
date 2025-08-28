import os
import discord
from discord import app_commands
from discord.ext import commands
from discord.ui import View, Button
import psycopg2
from typing import Optional
import io
import openpyxl
from discord import ui, Interaction

COMMITTEE_ROLES = {
    "Campus and Community Connections Committee": "🌐",
    "Technological Advancements Committee": "💻",
    "Graduate Affairs Committee": "🎓",
    "Academics and Research Committee": "📚",
}


class RoleButton(discord.ui.Button):
    def __init__(self, role_name: str, emoji: str):
        super().__init__(
            label=None,  # emoji-only button
            style=discord.ButtonStyle.secondary,
            custom_id=role_name,  # stable across restarts
            emoji=emoji
        )

    async def callback(self, interaction: Interaction):
        guild = interaction.guild
        if guild is None:
            return await interaction.response.send_message("Not in a server.", ephemeral=True)

        role = discord.utils.get(guild.roles, name=self.custom_id)
        if not role:
            return await interaction.response.send_message(
                f"⚠️ Role **{self.custom_id}** does not exist. Ask an admin to create it.", ephemeral=True
            )

        # Optional but helpful guards to avoid opaque 403s
        bot_member = guild.me
        if role.managed or role.is_default():
            return await interaction.response.send_message("I can't assign that role.", ephemeral=True)
        if bot_member.top_role <= role:
            return await interaction.response.send_message(
                "My top role is below that role. Move my role **above** committee roles.", ephemeral=True
            )

        member = interaction.user  # should be a Member in guild interactions
        try:
            if role in member.roles:
                await member.remove_roles(role, reason="Self-unassign")
                await interaction.response.send_message(f"❌ Removed **{role.name}**.", ephemeral=True)
            else:
                await member.add_roles(role, reason="Self-assign")
                await interaction.response.send_message(f"✅ Added **{role.name}**.", ephemeral=True)
        except discord.Forbidden:
            await interaction.response.send_message(
                "Missing access to modify roles. Check **Manage Roles** & role order.", ephemeral=True
            )


class RoleView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)
        for role_name, emoji in COMMITTEE_ROLES.items():
            self.add_item(RoleButton(role_name, emoji))


# ------------------------------------------------------------------------
# Environment variables (set these in .env, which docker-compose will load):
#
# DISCORD_TOKEN=YOUR_DISCORD_TOKEN
# DB_HOST=db           # "db" is the service name in docker-compose
# DB_USER=botuser
# DB_PASS=botpass
# DB_NAME=points_db
# GUILD_ID=you need to right click the ai club server and click copy id with developer mode on
# ------------------------------------------------------------------------

DISCORD_TOKEN = os.environ.get("DISCORD_TOKEN")

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASS = os.environ.get("DB_PASS", "")
DB_NAME = os.environ.get("DB_NAME", "points_db")

raw_gid = os.environ.get("GUILD_ID")
GUILD_ID: Optional[int] = int(
    raw_gid) if raw_gid and raw_gid.isdigit() else None

# 1) Connect to Postgres
conn = psycopg2.connect(
    host=DB_HOST,
    user=DB_USER,
    password=DB_PASS,
    dbname=DB_NAME
)
conn.autocommit = True  # So we don't have to manually commit
cursor = conn.cursor()

# 2) Ensure the necessary tables exist (if you haven't used init.sql)
cursor.execute("""
CREATE TABLE IF NOT EXISTS points (
    id SERIAL PRIMARY KEY,
    member_name VARCHAR(50) UNIQUE NOT NULL,
    points FLOAT DEFAULT 0
);

CREATE TABLE IF NOT EXISTS history (
    id SERIAL PRIMARY KEY,
    member_name VARCHAR(50) NOT NULL,
    reason TEXT,
    points FLOAT,  -- We'll store the delta (can be +x or -x)
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
);
""")

intents = discord.Intents.default()
intents.message_content = True

bot = commands.Bot(command_prefix="!", intents=intents)


PURGE_COMMANDS = os.environ.get("PURGE_COMMANDS") == "1"  # default off


@bot.event
async def on_ready():
    print(f"We have logged in as {bot.user}")
    try:
        guild = discord.Object(id=GUILD_ID) if GUILD_ID else None

        if guild:
            # Make guild-scoped copies of your global commands (instant visibility)
            bot.tree.copy_global_to(guild=guild)
            synced = await bot.tree.sync(guild=guild)
            print(f"Synced to guild {GUILD_ID}: {[c.name for c in synced]}")
        else:
            synced = await bot.tree.sync()
            print(f"Synced globally: {[c.name for c in synced]}")

        # Debug: what do we have locally vs remotely?
        print("Local commands:", [
              c.qualified_name for c in bot.tree.get_commands()])
        remote = await bot.tree.fetch_commands(guild=guild) if guild else await bot.tree.fetch_commands()
        print("Remote commands now:", [c.name for c in remote])

    except Exception as e:
        print(f"Error syncing commands: {e}")

    bot.add_view(RoleView())
    print(" Registered RoleView as persistent")


def _has_cabinet_role(interaction: discord.Interaction) -> bool:
    if interaction.guild is None:
        return False
    member = interaction.user  # discord.Member during guild interactions
    roles = getattr(member, "roles", []) or []
    return any("cabinet" in (r.name or "").lower() for r in roles)


def cabinet_only():
    """Use as @cabinet_only() on sensitive commands."""
    return app_commands.check(_has_cabinet_role)


async def _deny_ephemeral(interaction: discord.Interaction, msg: str = "You need a Cabinet role to run this command."):
    # Avoid double-responding if something else already did
    if interaction.response.is_done():
        try:
            await interaction.followup.send(msg, ephemeral=True)
        except Exception:
            pass
    else:
        await interaction.response.send_message(msg, ephemeral=True)


async def _send_codeblock_chunks(
    interaction: discord.Interaction,
    lines: list[str],
    *,
    ephemeral: bool = False,
):
    """Send a list of lines inside one or more ``` code blocks without exceeding 2000 chars.
       It will use interaction.response for the first send (if available), otherwise followups."""
    CHUNK_SIZE = 1900  # headroom for ``` and newlines
    chunks, buf = [], "```\n"
    for line in lines:
        if len(buf) + len(line) + 1 > CHUNK_SIZE:
            buf += "```"
            chunks.append(buf)
            buf = "```\n"
        buf += line + "\n"
    if buf != "```\n":
        buf += "```"
        chunks.append(buf)

    # choose the correct sender based on whether we've responded already
    first_send = (
        interaction.response.send_message
        if not interaction.response.is_done()
        else interaction.followup.send
    )

    await first_send(chunks[0], ephemeral=ephemeral)
    for chunk in chunks[1:]:
        await interaction.followup.send(chunk, ephemeral=ephemeral)


# --- ONE global error handler for all app commands ---
@bot.tree.error
async def on_app_command_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    # Permission (check) failures → quiet, friendly reply
    if isinstance(error, app_commands.CheckFailure):
        await _deny_ephemeral(interaction, "You need a Cabinet role to run this command.")
        return

    # Autocomplete errors or other expected issues you want to hide from users
    if isinstance(error, app_commands.CommandInvokeError):
        # You can introspect error.original if you want to branch
        pass

    # Fallback: log and show a generic message without leaking internals
    try:
        print(f"[appcmd error] {type(error).__name__}: {error}")
        if not interaction.response.is_done():
            await interaction.response.send_message("Something went wrong running that command.", ephemeral=True)
        else:
            await interaction.followup.send("Something went wrong running that command.", ephemeral=True)
    except Exception:
        # Avoid crashing the handler
        pass


async def member_autocomplete(
    interaction: discord.Interaction,
    current: str
) -> list[app_commands.Choice[str]]:
    """
    Return a list of up to 25 matching member names based on `current` partial input.
    """
    if not _has_cabinet_role(interaction):
        return []
    # For example, fetch from DB where member_name ILIKE '%current%'
    cursor.execute("""
        SELECT member_name
        FROM points
        WHERE member_name ILIKE %s
        ORDER BY member_name
        LIMIT 25
    """, (f"%{current}%",))
    rows = cursor.fetchall()

    return [
        app_commands.Choice(name=r[0], value=r[0]) for r in rows
    ]


@cabinet_only()
@bot.tree.command(name="addmember", description="Add a new member to the database with 0 starting points.")
@app_commands.describe(member_name="The name/surname of the member to add.")
async def addmember(interaction: discord.Interaction, member_name: str):
    """
    Adds a new member to the 'points' table with 0 points (if not existing).
    """
    cursor.execute(
        "SELECT member_name FROM points WHERE member_name = %s", (member_name,))
    row = cursor.fetchone()
    if row is not None:
        await interaction.response.send_message(
            f"Member '{member_name}' already exists in the database!",
            ephemeral=True
        )
        return

    cursor.execute(
        "INSERT INTO points (member_name, points) VALUES (%s, %s)",
        (member_name, 0.0)
    )
    await interaction.response.send_message(
        f"Member '{member_name}' added with 0 points."
    )


@cabinet_only()
@bot.tree.command(name="removemember", description="Remove a member and their history from the database.")
@app_commands.describe(member="The name/surname of the member to remove from the database.")
@app_commands.autocomplete(member=member_autocomplete)
async def removemember(interaction: discord.Interaction, member: str):
    """
    Removes the specified member from the 'points' table
    and deletes all corresponding logs from the 'history' table.
    """
    # 1) Check if member exists
    cursor.execute(
        "SELECT member_name FROM points WHERE member_name = %s", (member,))
    row = cursor.fetchone()
    if row is None:
        await interaction.response.send_message(
            f"Member '{member}' does not exist in the database!",
            ephemeral=True
        )
        return

    # 2) Delete rows from 'history' first (optional but usually desired)
    cursor.execute("DELETE FROM history WHERE member_name = %s", (member,))

    # 3) Remove from 'points'
    cursor.execute("DELETE FROM points WHERE member_name = %s", (member,))

    # 4) Confirmation
    await interaction.response.send_message(
        f"Member '{member}' has been removed from the database (including all history)."
    )


@cabinet_only()
@bot.tree.command(name="addpoints", description="Add points to an existing member, with a reason.")
@app_commands.describe(
    member="The member to award points to",
    amount="Number of points to add",
    reason="Reason for awarding the points"
)
@app_commands.autocomplete(member=member_autocomplete)
async def addpoints(interaction: discord.Interaction, member: str, amount: float, reason: str):
    """
    Adds points to an existing member and records the change in the 'history' table.
    """
    cursor.execute(
        "SELECT points FROM points WHERE member_name = %s", (member,))
    row = cursor.fetchone()
    if not row:
        await interaction.response.send_message(
            f"Member '{member}' does not exist. Use /addmember first!",
            ephemeral=True
        )
        return

    old_points = row[0]
    new_points = old_points + amount

    cursor.execute(
        "UPDATE points SET points = %s WHERE member_name = %s",
        (new_points, member)
    )
    cursor.execute(
        "INSERT INTO history (member_name, reason, points) VALUES (%s, %s, %s)",
        (member, reason, amount)
    )

    await interaction.response.send_message(
        f"**{amount}** points have been added to **{member}** for: *{reason}*\n"
        f"New total: **{new_points}** points."
    )


@cabinet_only()
@bot.tree.command(name="removepoints", description="Remove points from an existing member, with a reason.")
@app_commands.describe(
    member="The member to remove points from",
    amount="Number of points to remove",
    reason="Reason for removing the points"
)
@app_commands.autocomplete(member=member_autocomplete)  # <--- attach here
async def removepoints(interaction: discord.Interaction, member: str, amount: float, reason: str):
    """
    Subtracts points from an existing member and records the change in the 'history' table.
    """
    cursor.execute(
        "SELECT points FROM points WHERE member_name = %s", (member,))
    row = cursor.fetchone()
    if not row:
        await interaction.response.send_message(
            f"Member '{member}' does not exist. Use /addmember first!",
            ephemeral=True
        )
        return

    old_points = row[0]
    new_points = old_points - amount

    cursor.execute(
        "UPDATE points SET points = %s WHERE member_name = %s",
        (new_points, member)
    )
    cursor.execute(
        "INSERT INTO history (member_name, reason, points) VALUES (%s, %s, %s)",
        (member, reason, -amount)
    )

    await interaction.response.send_message(
        f"**{amount}** points have been removed from **{member}** for: *{reason}*\n"
        f"New total: **{new_points}** points."
    )


@bot.tree.command(name="showpoints", description="Show a particular member's points or everyone's.")
@app_commands.describe(member="Optionally specify a member to show points for. If omitted, shows all.")
@app_commands.autocomplete(member=member_autocomplete)
async def showpoints(interaction: discord.Interaction, member: Optional[str] = None):
    """
    If 'member' is provided, show points for that one member.
    Otherwise, show points for everyone (paged to respect Discord 2000-char limit).
    """
    if member is None:
        cursor.execute(
            "SELECT member_name, points FROM points ORDER BY member_name ASC")
        rows = cursor.fetchall()
        if not rows:
            await interaction.response.send_message("No members in the database yet!")
            return

        # format a simple table
        header = ["Name", "Points"]
        sep = ["------------", "------"]
        lines = [f"{header[0]:<28}  {header[1]}",
                 f"{sep[0]:<28}  {sep[1]}"]
        for name_, pts in rows:
            # use :g to avoid trailing .0 spam; switch to :.2f if you prefer fixed decimals
            lines.append(f"{name_:<28}  {pts:g}")

        await _send_codeblock_chunks(interaction, lines)
        return

    # specific member
    cursor.execute(
        "SELECT member_name, points FROM points WHERE member_name = %s", (member,))
    row = cursor.fetchone()
    if not row:
        await interaction.response.send_message(f"Member '{member}' does not exist in the database.", ephemeral=True)
        return

    member_name, pts = row
    lines = [
        f"{'Name':<28}  Points",
        f"{'------------':<28}  ------",
        f"{member_name:<28}  {pts:g}",
    ]
    await _send_codeblock_chunks(interaction, lines, ephemeral=False)


@cabinet_only()
@bot.tree.command(name="showlogs", description="Show the historical logs for a given member.")
@app_commands.describe(member="The member whose history/logs you want to see.")
# <--- attach here as well
@app_commands.autocomplete(member=member_autocomplete)
async def showlogs(interaction: discord.Interaction, member: str):
    """
    Shows all history logs for a specific member, sorted by the creation time.
    """
    cursor.execute(
        "SELECT member_name FROM points WHERE member_name = %s", (member,))
    row = cursor.fetchone()
    if not row:
        await interaction.response.send_message(
            f"Member '{member}' does not exist. Use /addmember first!",
            ephemeral=True
        )
        return

    cursor.execute("""
        SELECT reason, points, created_at
        FROM history
        WHERE member_name = %s
        ORDER BY created_at ASC
    """, (member,))

    logs = cursor.fetchall()
    if not logs:
        await interaction.response.send_message(
            f"No history logs found for member '{member}'."
        )
        return

    # Build a table of logs
    logs_table = "```\nReason                           Points   Timestamp\n"
    logs_table += "--------------------------------  -------  ---------------------\n"

    for reason, points, ts in logs:
        reason_str = (reason[:30] + "...") if len(reason) > 30 else reason
        logs_table += f"{reason_str:32}  {points:7}  {ts}\n"

    logs_table += "```"
    await interaction.response.send_message(logs_table)


@cabinet_only()
@bot.tree.command(name="showmembers", description="Show a list of all members in the database.")
async def showmembers(interaction: discord.Interaction):
    """
    Shows a list of all members in the database, paged to respect Discord's 2000-char limit.
    """
    cursor.execute("SELECT member_name FROM points ORDER BY member_name ASC")
    rows = cursor.fetchall()
    if not rows:
        await interaction.response.send_message("No members in the database yet!")
        return

    members = [r[0] for r in rows]
    lines = ["Members", "------------"] + members
    await _send_codeblock_chunks(interaction, lines)


@cabinet_only()
@bot.tree.command(
    name="addmembers_fromexcel",
    description="Upload an Excel file with First/Last Name columns to add members."
)
@app_commands.describe(file="Excel file (.xlsx) with First Name and Last Name columns")
async def addmembers_fromexcel(interaction: discord.Interaction, file: discord.Attachment):
    # Check file type
    if not file.filename.lower().endswith(".xlsx"):
        await interaction.response.send_message("Please upload a valid .xlsx Excel file.", ephemeral=True)
        return

    file_bytes = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active  # first sheet

    # --- Robust header detection (nil-safe + flexible names) ---
    import re

    def norm(s: Optional[str]) -> str:
        # lower, trim, remove non-alphanumerics to normalize variations like "First name", "first-name", etc.
        return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower())

    # Read header row as plain values (no .value access needed)
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    norm_to_index0 = {norm(v): i for i, v in enumerate(
        header_row)}  # 0-based indices

    # Accept a few common variants
    first_candidates = {"firstname", "first", "fname"}
    last_candidates = {"lastname", "last", "lname", "surname", "familyname"}

    def find_index(candidates: set[str]) -> Optional[int]:
        for key in candidates:
            if key in norm_to_index0:
                return norm_to_index0[key]
        return None

    first_i = find_index(first_candidates)
    last_i = find_index(last_candidates)

    if first_i is None or last_i is None:
        pretty = ", ".join(
            str(v) if v is not None else "∅" for v in header_row)
        await interaction.response.send_message(
            "Excel must contain **First Name** and **Last Name** columns.\n"
            "Accepted header variants:\n"
            " • First Name: First name / First / FirstName / FName\n"
            " • Last Name:  Last name / Last / LastName / LName / Surname / Family Name\n"
            f"Detected header row: {pretty}",
            ephemeral=True
        )
        return

    # --- Ingest rows ---
    added, skipped = [], []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        first = (row[first_i] or "").strip()
        last = (row[last_i] or "").strip()
        if not first or not last:
            continue

        member_name = f"{first} {last}"
        # Up to you: enforce a simple whitelist to avoid weird DB entries
        # if not re.fullmatch(r"[A-Za-z][A-Za-z '\-]{0,48}[A-Za-z]?", member_name):
        #     continue

        # Insert if not exists
        cursor.execute(
            "SELECT 1 FROM points WHERE member_name = %s", (member_name,))
        if cursor.fetchone() is None:
            cursor.execute(
                "INSERT INTO points (member_name, points) VALUES (%s, %s)",
                (member_name, 0.0)
            )
            added.append(member_name)
        else:
            skipped.append(member_name)

    # --- Report ---
    msg = f"✅ Added {len(added)} members."
    if skipped:
        msg += f"  ⚠️ Skipped {len(skipped)} duplicate{'s' if len(skipped) != 1 else ''} (already existed)."
    if added:
        msg += "\nNewly added: " + \
            ", ".join(added[:10]) + ("..." if len(added) > 10 else "")

    # Send summary first
    await interaction.response.send_message(msg, ephemeral=True)

    # If there are duplicates, send their names as a paged code block list (now safe)
    if skipped:
        lines = ["Duplicates (already existed)",
                 "-----------------------------"] + skipped
        await _send_codeblock_chunks(interaction, lines, ephemeral=True)


@cabinet_only()
@bot.tree.command(name="renamemember", description="Rename a member and update their history.")
@app_commands.describe(
    old_member="Existing member to rename",
    new_name="New name (e.g., 'First Last')"
)
@app_commands.autocomplete(old_member=member_autocomplete)
async def renamemember(interaction: discord.Interaction, old_member: str, new_name: str):
    # --- normalize new name ---
    new_name = " ".join((new_name or "").strip().split()
                        )  # collapse extra spaces
    if not new_name:
        await interaction.response.send_message("New name cannot be empty.", ephemeral=True)
        return
    if len(new_name) > 50:
        await interaction.response.send_message("New name must be ≤ 50 characters.", ephemeral=True)
        return

    # --- ensure old exists ---
    cursor.execute(
        "SELECT points FROM points WHERE member_name = %s", (old_member,))
    row = cursor.fetchone()
    if not row:
        await interaction.response.send_message(
            f"Member '{old_member}' does not exist.", ephemeral=True
        )
        return

    # --- block duplicates ---
    cursor.execute("SELECT 1 FROM points WHERE member_name = %s", (new_name,))
    if cursor.fetchone():
        await interaction.response.send_message(
            f"Cannot rename to '{new_name}' because that name already exists.", ephemeral=True
        )
        return

    # --- do the rename in both tables ---
    # points table (unique key lives here)
    cursor.execute(
        "UPDATE points SET member_name = %s WHERE member_name = %s",
        (new_name, old_member)
    )
    # history table (non-unique, update all rows)
    cursor.execute(
        "UPDATE history SET member_name = %s WHERE member_name = %s",
        (new_name, old_member)
    )

    await interaction.response.send_message(
        f"✅ Renamed **{old_member}** → **{new_name}**."
    )


@bot.tree.command(name="membercount", description="Show the total number of members in the AI Club.")
async def membercount(interaction: discord.Interaction):
    cursor.execute("SELECT COUNT(*) FROM points")
    (count,) = cursor.fetchone()
    await interaction.response.send_message(f"Total members: **{count}**")


@cabinet_only()
@bot.tree.command(name="removeallmembers", description="⚠️ Remove ALL members and their history from the database.")
async def removeallmembers(interaction: discord.Interaction):
    """Ask for confirmation before removing all members."""

    class ConfirmView(View):
        def __init__(self, owner_id: int):
            super().__init__(timeout=30)  # 30s timeout
            self.owner_id = owner_id

        async def interaction_check(self, i: discord.Interaction) -> bool:
            # Only the original invoker can press buttons
            if i.user.id != self.owner_id:
                await i.response.send_message("You can't interact with this confirmation.", ephemeral=True)
                return False
            return True

        async def on_timeout(self):
            # Disable buttons when timing out
            for item in self.children:
                if isinstance(item, Button):
                    item.disabled = True
            try:
                await self.message.edit(content="⏳ Confirmation timed out. No changes made.", view=self)
            except Exception:
                pass

        @discord.ui.button(label="YES, DELETE EVERYTHING", style=discord.ButtonStyle.danger)
        async def confirm(self, interaction_btn: discord.Interaction, button: Button):
            # Perform deletion
            cursor.execute("DELETE FROM history")
            cursor.execute("DELETE FROM points")

            # Disable buttons after action
            for item in self.children:
                if isinstance(item, Button):
                    item.disabled = True

            await interaction_btn.response.edit_message(
                content="⚠️ All members and their history have been **permanently removed**.",
                view=self
            )

        @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary)
        async def cancel(self, interaction_btn: discord.Interaction, button: Button):
            for item in self.children:
                if isinstance(item, Button):
                    item.disabled = True
            await interaction_btn.response.edit_message(
                content="Action cancelled. Database unchanged.",
                view=self
            )

    view = ConfirmView(owner_id=interaction.user.id)
    # Keep a handle to edit on timeout
    await interaction.response.send_message(
        "⚠️ **Are you absolutely sure?** This will permanently delete **all members and their history**.",
        view=view,
        ephemeral=True
    )
    # Save the message so on_timeout can edit it
    view.message = await interaction.original_response()


@cabinet_only()
@bot.tree.command(name="showroles", description="Show committee roles you can self-assign")
async def showroles(interaction: discord.Interaction):
    """
    Shows buttons for users to self-assign/remove committee roles.
    """
    await interaction.response.send_message(
        "📌 Select the committee(s) you want to join by clicking the buttons below:\n\nCampus and Community Connections Committee: 🌐,\nTechnological Advancements Committee: 💻,\nGraduateAffairs Committee: 🎓,\nAcademics and Research Committee: 📚 \n \u200B",
        view=RoleView()
    )

bot.run(DISCORD_TOKEN)
